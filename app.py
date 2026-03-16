import io
import os
import re
from itertools import combinations
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

st.set_page_config(
    page_title="Recharge Rate Analyzer",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =========================================================
# LIGHT THEME CSS
# =========================================================
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Syne:wght@400;600;700;800&display=swap');

  html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: #ffffff;
    color: #1e293b;
  }
  .stApp {
    background: #f8fafc;
  }
  header[data-testid="stHeader"] {
    background: transparent;
  }

  /* Title */
  .main-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.2rem;
    font-weight: 800;
    color: #0f172a;
    margin-bottom: 0.2rem;
  }
  .main-title span {
    background: linear-gradient(90deg, #2563eb, #7c3aed);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
  }
  .main-subtitle {
    color: #94a3b8;
    font-size: 0.85rem;
    font-family: 'JetBrains Mono', monospace;
    margin-bottom: 2rem;
  }

  /* Metric cards */
  .metric-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    position: relative;
    overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }
  .metric-card::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #2563eb, #7c3aed);
  }
  .metric-label {
    font-size: 0.72rem;
    font-family: 'JetBrains Mono', monospace;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.4rem;
  }
  .metric-value {
    font-size: 1.8rem;
    font-weight: 800;
    color: #0f172a;
    line-height: 1;
  }
  .metric-sub {
    font-size: 0.72rem;
    color: #2563eb;
    margin-top: 0.35rem;
    font-family: 'JetBrains Mono', monospace;
  }
  .metric-green .metric-value { color: #16a34a; }
  .metric-green .metric-sub   { color: #16a34a; }

  /* Opportunity cards */
  .opp-card {
    background: #f0fdf4;
    border: 1px solid #bbf7d0;
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    position: relative;
    overflow: hidden;
    height: 100%;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
  }
  .opp-card::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #16a34a, #4ade80);
  }
  .opp-rank {
    font-size: 0.7rem;
    font-family: 'JetBrains Mono', monospace;
    color: #15803d;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.4rem;
  }
  .opp-profit {
    font-size: 1.9rem;
    font-weight: 800;
    color: #16a34a;
    font-family: 'JetBrains Mono', monospace;
    line-height: 1;
    margin-bottom: 0.5rem;
  }
  .opp-detail {
    font-size: 0.82rem;
    color: #475569;
    line-height: 1.7;
  }

  /* Section headers */
  .section-header {
    font-family: 'Syne', sans-serif;
    font-size: 1.05rem;
    font-weight: 700;
    color: #0f172a;
    border-left: 3px solid #2563eb;
    padding-left: 0.75rem;
    margin: 2rem 0 0.9rem 0;
  }

  /* Upload area */
  [data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 2px dashed #cbd5e1 !important;
    border-radius: 14px !important;
  }
  [data-testid="stFileUploader"]:hover {
    border-color: #2563eb !important;
  }

  /* Dataframe */
  [data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid #e2e8f0 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
  }

  /* Tabs */
  .stTabs [data-baseweb="tab-list"] {
    background: #f1f5f9;
    border-radius: 10px;
    padding: 3px;
    gap: 3px;
  }
  .stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 7px;
    color: #64748b;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.8rem;
  }
  .stTabs [aria-selected="true"] {
    background: #ffffff !important;
    color: #2563eb !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }

  /* Download buttons */
  .stDownloadButton button {
    background: #ffffff !important;
    color: #2563eb !important;
    border: 1px solid #bfdbfe !important;
    border-radius: 10px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.78rem !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06) !important;
  }
  .stDownloadButton button:hover {
    background: #eff6ff !important;
    border-color: #2563eb !important;
  }

  /* Number inputs */
  [data-testid="stNumberInput"] input {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    color: #0f172a !important;
    border-radius: 8px !important;
    font-family: 'JetBrains Mono', monospace !important;
  }

  /* Expander */
  .streamlit-expanderHeader {
    background: #f8fafc !important;
    border-radius: 10px !important;
    color: #64748b !important;
    font-size: 0.82rem !important;
    border: 1px solid #e2e8f0 !important;
  }

  /* Alerts */
  .stSuccess > div {
    background: #f0fdf4 !important;
    border: 1px solid #bbf7d0 !important;
    border-radius: 10px !important;
    color: #15803d !important;
  }
  .stInfo > div {
    background: #eff6ff !important;
    border: 1px solid #bfdbfe !important;
    border-radius: 10px !important;
    color: #1d4ed8 !important;
  }
  .stWarning > div {
    background: #fffbeb !important;
    border: 1px solid #fde68a !important;
    border-radius: 10px !important;
    color: #92400e !important;
  }

  hr { border-color: #e2e8f0 !important; }
</style>
""", unsafe_allow_html=True)


# =========================================================
# HELPERS
# =========================================================
def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_rate(value) -> Optional[float]:
    if value is None:
        return None
    try:
        s = str(value).replace("%", "").strip()
        num = float(s)
    except Exception:
        return None
    if num < 0:
        return None
    return num


def parse_amount_spec(value: Any) -> Optional[Dict[str, Any]]:
    """Parse any amount cell into a dict with type and a pre-computed set of amounts."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    s = str(value).strip().replace("₹", "")
    s = re.sub(r"\s+", "", s)
    if not s:
        return None
    if s.upper() in {"ANY", "ALL"}:
        return {"type": "all", "values": None, "_set": None}

    result_ints: List[int] = []
    parts = [p.strip() for p in re.split(r"[,，]", s) if p.strip()]
    for part in parts:
        part = re.sub(r"\s+", "", part)
        if not part:
            continue
        m_range = re.fullmatch(r"(\d+)-(\d+)", part)
        if m_range:
            a, b = int(m_range.group(1)), int(m_range.group(2))
            result_ints.extend(range(min(a, b), max(a, b) + 1))
            continue
        m_single = re.fullmatch(r"(\d+)", part)
        if m_single:
            result_ints.append(int(m_single.group(1)))
    if not result_ints:
        return None
    s2 = set(result_ints)
    return {"type": "list", "values": sorted(s2), "_set": s2}


def rule_applies(rule: Dict[str, Any], amount: int, _cache: dict = {}) -> bool:
    if rule["type"] == "all":
        return True
    return amount in rule["_set"]


# =========================================================
# ZONE NORMALISATION
# =========================================================
_ZONE_ALIASES: Dict[str, str] = {
    # UP East variants
    "up east": "UP East",
    "uttar pradesh (e)": "UP East",
    "uttar pradesh east": "UP East",
    "upe": "UP East",
    # UP West variants
    "up west": "UP West",
    "uttar pradesh (w)": "UP West",
    "uttar pradesh west": "UP West",
    "upw": "UP West",
    # Add more as needed
}

def normalize_zone(raw: str) -> str:
    key = raw.strip().lower()
    return _ZONE_ALIASES.get(key, raw.strip())


# =========================================================
# EXCEL PARSING  (format-agnostic)
# =========================================================
def find_header_row(ws) -> Optional[int]:
    expected = {"operator", "zone", "amount", "amounts", "margin"}
    max_scan = min(ws.max_row, 20)
    for r in range(1, max_scan + 1):
        row_vals = [clean_text(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if sum(1 for v in row_vals if v in expected) >= 2:
            return r
    return None


def find_target_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        if find_header_row(ws) is not None:
            return ws
    return wb[wb.sheetnames[0]]


def get_column_map(ws, header_row: int) -> Dict[str, int]:
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = clean_text(ws.cell(header_row, c).value).lower()
        if val:
            col_map[val] = c
    return col_map


def load_supplier_rows(uploaded_file, supplier_name: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
    ws = find_target_sheet(wb)
    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError(f"Could not locate header row in {uploaded_file.name}")
    col_map = get_column_map(ws, header_row)

    operator_col = col_map.get("operator")
    zone_col = col_map.get("zone")
    amount_col = col_map.get("amount") or col_map.get("amounts")
    if amount_col is None:
        for k, v in col_map.items():
            if "amount" in k:
                amount_col = v
                break
    margin_col = col_map.get("margin")

    if not operator_col or not amount_col or not margin_col:
        raise ValueError(
            f"{uploaded_file.name}: could not find Operator, Amount, and Margin columns. "
            f"Found: {list(col_map.keys())}"
        )

    # ── Per-file margin format detection ──────────────────────────────────
    # Scan all margin cells to decide whether they are stored as:
    #   (A) already-percent  e.g. 2.41 meaning 2.41%   → use as-is
    #   (B) decimal fraction e.g. 0.0268 meaning 2.68% → multiply × 100
    # If the Excel cell format contains "%" we always use (B).
    # For General-formatted cells, we use majority vote: if most values > 1 → (A).
    _fmt_sample: List[str] = []
    _val_sample: List[float] = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, margin_col).value
        f = ws.cell(r, margin_col).number_format or ""
        if v is not None:
            try:
                _fmt_sample.append(f)
                _val_sample.append(float(v))
            except Exception:
                pass

    _has_pct_fmt = any("%" in f for f in _fmt_sample)
    if _has_pct_fmt:
        # RC-style: stored as decimal fraction (0.0268) with % cell format
        _margin_mode = "decimal"
    else:
        # PD-style: count how many values are already > 1
        _above_one = sum(1 for v in _val_sample if v > 1)
        _margin_mode = "percent" if _above_one >= len(_val_sample) / 2 else "decimal"
    # ──────────────────────────────────────────────────────────────────────

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        operator_val = ws.cell(r, operator_col).value
        zone_val = ws.cell(r, zone_col).value if zone_col else "ALL Zone"
        amount_val = ws.cell(r, amount_col).value
        margin_val = ws.cell(r, margin_col).value

        if operator_val is None or amount_val is None or margin_val is None:
            continue

        try:
            fv = float(str(margin_val).replace("%", "").strip())
        except Exception:
            continue

        if _margin_mode == "decimal":
            rate = round(fv, 4)
        else:
            rate = round(fv, 4)

        if rate <= 0:
            continue

        rule = parse_amount_spec(amount_val)
        if rule is None:
            continue

        rows.append({
            "supplier": supplier_name,
            "operator": clean_text(operator_val),
            "zone": normalize_zone(clean_text(zone_val) or "ALL Zone"),
            "rule": rule,
            "raw_rule": clean_text(amount_val),
            "rate": rate,
            "excel_row": r,
            "file_name": uploaded_file.name,
        })
    return rows


# =========================================================
# BEST RATE MAP
# =========================================================
def build_best_rate_map(rows, min_amt, max_amt):
    best = {}
    for amt in range(min_amt, max_amt + 1):
        best_row, best_rate = None, -1.0
        for row in rows:
            rule = row["rule"]
            applies = (rule["type"] == "all") or (amt in rule["_set"])
            if applies and row["rate"] > best_rate:
                best_rate = row["rate"]
                best_row = row
        if best_row:
            best[amt] = best_row
    return best


def compress_best_map(best_map):
    if not best_map:
        return []
    amounts = sorted(best_map.keys())
    segments, start, prev, current = [], amounts[0], amounts[0], best_map[amounts[0]]
    for amt in amounts[1:]:
        row = best_map[amt]
        same = (amt == prev + 1 and row["supplier"] == current["supplier"]
                and row["operator"] == current["operator"] and row["zone"] == current["zone"]
                and row["rate"] == current["rate"] and row["raw_rule"] == current["raw_rule"])
        if same:
            prev = amt
            continue
        segments.append({"supplier": current["supplier"], "operator": current["operator"],
                          "zone": current["zone"], "amount_from": start, "amount_to": prev,
                          "rate_percent": round(current["rate"], 4), "source_rule": current["raw_rule"]})
        start, prev, current = amt, amt, row
    segments.append({"supplier": current["supplier"], "operator": current["operator"],
                      "zone": current["zone"], "amount_from": start, "amount_to": prev,
                      "rate_percent": round(current["rate"], 4), "source_rule": current["raw_rule"]})
    return segments


# =========================================================
# DIAGNOSTICS
# =========================================================
def find_coverage_gaps(rows, min_amt, max_amt):
    covered = set()
    for amt in range(min_amt, max_amt + 1):
        for row in rows:
            rule = row["rule"]
            if rule["type"] == "all" or amt in rule["_set"]:
                covered.add(amt)
                break
    gaps, gap_start = [], None
    for amt in range(min_amt, max_amt + 1):
        if amt not in covered:
            if gap_start is None: gap_start = amt
        elif gap_start is not None:
            gaps.append((gap_start, amt - 1))
            gap_start = None
    if gap_start is not None:
        gaps.append((gap_start, max_amt))
    if not rows: return []
    s = rows[0]
    return [{"supplier": s["supplier"], "operator": s["operator"], "zone": s["zone"],
             "missing_from": a, "missing_to": b} for a, b in gaps]


def find_overlaps(rows, min_amt, max_amt, limit=200):
    overlaps = []
    for amt in range(min_amt, max_amt + 1):
        matching = [r for r in rows if r["rule"]["type"] == "all" or amt in r["rule"]["_set"]]
        if len(matching) > 1:
            winning = max(matching, key=lambda x: x["rate"])
            overlaps.append({"supplier": winning["supplier"], "operator": winning["operator"],
                              "zone": winning["zone"], "amount": amt,
                              "matching_rules": ", ".join(sorted(set(r["raw_rule"] for r in matching))),
                              "winning_rule": winning["raw_rule"],
                              "winning_rate_percent": round(winning["rate"], 4)})
        if len(overlaps) >= limit: break
    return overlaps


def find_redundant_rules(rows, best_map):
    winning_keys = {(r["excel_row"], r["raw_rule"], r["rate"]) for r in best_map.values()}
    return [{"supplier": r["supplier"], "operator": r["operator"], "zone": r["zone"],
             "rule": r["raw_rule"], "rate_percent": round(r["rate"], 4), "excel_row": r["excel_row"],
             "reason": "Never highest for any denomination in selected range"}
            for r in rows if (r["excel_row"], r["raw_rule"], r["rate"]) not in winning_keys]


# =========================================================
# SUPPLIER COMPARISON
# =========================================================
def compare_suppliers(best_a, best_b, name_a, name_b):
    """Return one row per amount where the two suppliers have a different best rate.

    Amounts where BOTH suppliers are only matched by a catch-all rule (ANY / ALL)
    are skipped — those are phantom denominations that don't actually exist in
    the real recharge catalogue and produce junk opportunities.
    """
    opportunities = []
    for amt in sorted(set(best_a.keys()) & set(best_b.keys())):
        row_a, row_b = best_a[amt], best_b[amt]

        # # Skip if both sides are purely catch-all coverage
        # a_is_catchall = row_a["rule"]["type"] == "all"
        # b_is_catchall = row_b["rule"]["type"] == "all"
        # if a_is_catchall and b_is_catchall:
        #     continue

        r1, r2 = row_a["rate"], row_b["rate"]
        if r1 == r2:
            continue
        if r1 > r2:
            buy_from, sell_to, buy_rate, sell_rate = name_a, name_b, r1, r2
            buy_rule, sell_rule = row_a["raw_rule"], row_b["raw_rule"]
        else:
            buy_from, sell_to, buy_rate, sell_rate = name_b, name_a, r2, r1
            buy_rule, sell_rule = row_b["raw_rule"], row_a["raw_rule"]
        opportunities.append({
            "amount": amt,
            "operator": row_a["operator"],
            "zone": row_a["zone"],
            "buy_from": buy_from,
            "sell_to": sell_to,
            "buy_rate_percent": round(buy_rate, 4),
            "sell_rate_percent": round(sell_rate, 4),
            "profit_percent": round(buy_rate - sell_rate, 4),
            "buy_rule": buy_rule,
            "sell_rule": sell_rule,
        })
    return opportunities


def compress_opportunities(opportunities: List[Dict]) -> List[Dict]:
    """Merge consecutive amounts with identical parameters into amount_from / amount_to segments."""
    if not opportunities:
        return []
    segments = []
    start = prev = opportunities[0]["amount"]
    current = opportunities[0]
    for item in opportunities[1:]:
        same = (
            item["amount"] == prev + 1
            and item["operator"] == current["operator"]
            and item["zone"] == current["zone"]
            and item["buy_from"] == current["buy_from"]
            and item["sell_to"] == current["sell_to"]
            and item["buy_rate_percent"] == current["buy_rate_percent"]
            and item["sell_rate_percent"] == current["sell_rate_percent"]
        )
        if same:
            prev = item["amount"]
            continue
        segments.append({
            "operator": current["operator"], "zone": current["zone"],
            "amount_from": start, "amount_to": prev,
            "buy_from": current["buy_from"], "sell_to": current["sell_to"],
            "buy_rate_percent": current["buy_rate_percent"],
            "sell_rate_percent": current["sell_rate_percent"],
            "profit_percent": current["profit_percent"],
            "buy_rule": current["buy_rule"],
            "sell_rule": current["sell_rule"],
        })
        start = prev = item["amount"]
        current = item
    segments.append({
        "operator": current["operator"], "zone": current["zone"],
        "amount_from": start, "amount_to": prev,
        "buy_from": current["buy_from"], "sell_to": current["sell_to"],
        "buy_rate_percent": current["buy_rate_percent"],
        "sell_rate_percent": current["sell_rate_percent"],
        "profit_percent": current["profit_percent"],
        "buy_rule": current["buy_rule"],
        "sell_rule": current["sell_rule"],
    })
    return segments


# =========================================================
# DEAL CLUBBING
# =========================================================
def _fmt_slabs(rows_in_group: pd.DataFrame) -> str:
    """Format a list of amount_from/amount_to rows as compact slab strings."""
    parts = []
    for _, r in rows_in_group.sort_values("amount_from").iterrows():
        lo, hi = int(r["amount_from"]), int(r["amount_to"])
        parts.append(str(lo) if lo == hi else f"{lo}-{hi}")
    return ", ".join(parts)


def club_deals(opportunities_df: pd.DataFrame, margin_tolerance: float = 0.10) -> pd.DataFrame:
    """
    Group opportunity segments into portal-ready clubbed deals.

    Two segments are clubbed when ALL of these match:
      • operator, zone, buy_from, sell_to
      • buy_rate_percent  (identical — same rate you'd type when buying)
      • sell_rate_percent (identical — same rate you'd type when selling)
      • |profit_percent difference| ≤ margin_tolerance

    Because buy_rate and sell_rate are both fixed, profit is also identical
    (tolerance check is a safety net for any floating-point drift).

    Output sorted descending by profit_percent (highest GP deal first).
    """
    if opportunities_df.empty:
        return pd.DataFrame()

    df = opportunities_df.copy().reset_index(drop=True)
    group_key = ["operator", "zone", "buy_from", "sell_to",
                 "buy_rate_percent", "sell_rate_percent"]

    clubbed_rows = []
    for key_vals, grp in df.groupby(group_key, sort=False):
        # Within a same-rate group, further split if profit drifts beyond tolerance
        # (can happen at boundaries where one amount falls on a different source rule)
        grp = grp.sort_values("profit_percent", ascending=False).reset_index(drop=True)
        anchor_profit = grp.loc[0, "profit_percent"]
        within_tol = grp[abs(grp["profit_percent"] - anchor_profit) <= margin_tolerance]
        outside_tol = grp[abs(grp["profit_percent"] - anchor_profit) > margin_tolerance]

        def emit(sub: pd.DataFrame):
            if sub.empty:
                return
            row0 = sub.iloc[0]
            clubbed_rows.append({
                "operator":      row0["operator"],
                "zone":          row0["zone"],
                "buy_from":      row0["buy_from"],
                "sell_to":       row0["sell_to"],
                "slabs":         _fmt_slabs(sub),
                "slab_count":    len(sub),
                "buy_rate_%":    row0["buy_rate_percent"],
                "sell_rate_%":   row0["sell_rate_percent"],
                "profit_%":      row0["profit_percent"],
                "supplier_pair": sub["supplier_pair"].iloc[0] if "supplier_pair" in sub.columns else "",
            })

        emit(within_tol)
        # Recursively emit the remainder as separate deal(s)
        if not outside_tol.empty:
            emit(outside_tol)

    result = pd.DataFrame(clubbed_rows)
    if not result.empty:
        result = result.sort_values("profit_%", ascending=True).reset_index(drop=True)
        result.insert(0, "#", range(1, len(result) + 1))
    return result


# =========================================================
# EXPORT
# =========================================================
def to_excel_bytes(summary_df, best_df, opportunities_df, gaps_df, overlaps_df, redundant_df, clubbed_df=None):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets_written = 0
        for df, sheet in [(summary_df, "Summary"), (best_df, "Best_Rates"),
                          (opportunities_df, "Buy_Sell_Opportunities"), (gaps_df, "Coverage_Gaps"),
                          (overlaps_df, "Overlaps"), (redundant_df, "Redundant_Rules")]:
            if not df.empty:
                df.to_excel(writer, index=False, sheet_name=sheet)
                sheets_written += 1
        if clubbed_df is not None and not clubbed_df.empty:
            clubbed_df.to_excel(writer, index=False, sheet_name="Clubbed_Deals")
            sheets_written += 1
        if sheets_written == 0:
            # openpyxl requires at least one visible sheet
            pd.DataFrame({"info": ["No data — upload supplier files to generate analysis."]}).to_excel(
                writer, index=False, sheet_name="Info"
            )
    output.seek(0)
    return output.read()


# =========================================================
# CHART HELPERS
# =========================================================
CHART_THEME = dict(
    paper_bgcolor="#ffffff",
    plot_bgcolor="#f8fafc",
    font=dict(color="#475569", family="JetBrains Mono", size=11),
    xaxis=dict(gridcolor="#e2e8f0", zerolinecolor="#e2e8f0"),
    yaxis=dict(gridcolor="#e2e8f0", zerolinecolor="#e2e8f0"),
    margin=dict(l=20, r=20, t=44, b=20),
)

SUPPLIER_COLORS = ["#2563eb", "#7c3aed", "#059669", "#ea580c", "#db2777", "#ca8a04", "#0891b2", "#65a30d"]


def chart_rate_comparison(best_df, suppliers):
    if best_df.empty: return None
    agg = best_df.groupby(["supplier", "operator"])["rate_percent"].max().reset_index()
    color_map = {s: SUPPLIER_COLORS[i % len(SUPPLIER_COLORS)] for i, s in enumerate(suppliers)}
    fig = px.bar(agg, x="operator", y="rate_percent", color="supplier", barmode="group",
                 color_discrete_map=color_map,
                 labels={"rate_percent": "Best Rate (%)", "operator": "Operator", "supplier": "Supplier"},
                 title="Best Rate by Operator & Supplier")
    fig.update_layout(**CHART_THEME, title_font=dict(size=13, color="#0f172a"),
                      legend=dict(bgcolor="#f8fafc", bordercolor="#e2e8f0", font=dict(color="#475569")))
    fig.update_traces(marker_line_width=0)
    return fig


def chart_profit_bar(opp_df):
    if opp_df.empty: return None
    df = opp_df.copy()
    df["label"] = df["operator"] + "  ₹" + df["amount_from"].astype(str) + "–" + df["amount_to"].astype(str)
    fig = px.bar(df.sort_values("profit_percent", ascending=True),
                 x="profit_percent", y="label", orientation="h",
                 color="profit_percent",
                 color_continuous_scale=[[0, "#bbf7d0"], [1, "#15803d"]],
                 labels={"profit_percent": "Profit (%)", "label": ""},
                 title="Profit by Opportunity")
    fig.update_layout(**CHART_THEME, title_font=dict(size=13, color="#0f172a"), coloraxis_showscale=False)
    fig.update_traces(marker_line_width=0)
    return fig


def chart_heatmap(best_df):
    if best_df.empty or best_df["supplier"].nunique() < 2: return None
    pivot = best_df.groupby(["supplier", "operator"])["rate_percent"].max().unstack(fill_value=0)
    fig = go.Figure(data=go.Heatmap(
        z=pivot.values, x=pivot.columns.tolist(), y=pivot.index.tolist(),
        colorscale=[[0, "#eff6ff"], [0.5, "#93c5fd"], [1, "#1d4ed8"]],
        text=pivot.values.round(4), texttemplate="%{text}%",
        textfont={"size": 11, "color": "#0f172a"},
    ))
    fig.update_layout(**CHART_THEME, title="Rate Heatmap", title_font=dict(size=13, color="#0f172a"))
    return fig


# =========================================================
# UI — HEADER
# =========================================================
st.markdown('<div class="main-title">⚡ Recharge <span>Rate Analyzer</span></div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">multi-supplier margin intelligence</div>', unsafe_allow_html=True)

col_up, col_min, col_max = st.columns([3, 1, 1])
with col_up:
    uploaded_files = st.file_uploader(
        "Drop supplier Excel files here",
        type=["xlsx"], accept_multiple_files=True,
        help="Each file = one supplier. Needs Operator, Amount/Amounts, Margin columns.",
    )
with col_min:
    min_amt = int(st.number_input("Min Denomination", min_value=1, value=1, step=1))
with col_max:
    max_amt = int(st.number_input("Max Denomination", min_value=1, value=50000, step=1))

with st.expander("📋 Expected Excel format", expanded=False):
    st.markdown("Required columns: **Operator**, **Zone**, **Amount** or **Amounts**, **Margin**\n\nMargin is read as-is. `1.85` and `1.85%` both mean **1.85%**. Excel %-formatted cells are auto-converted.")

if not uploaded_files:
    st.markdown("""
    <div style="text-align:center;padding:4rem 2rem;color:#cbd5e1;">
        <div style="font-size:2.5rem;margin-bottom:0.8rem">📂</div>
        <div style="font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:#94a3b8">
            Upload supplier Excel files above to begin analysis
        </div>
    </div>""", unsafe_allow_html=True)
    st.stop()

# =========================================================
# PROCESSING
# =========================================================
supplier_maps, summary_rows = {}, []
all_best_segments, all_gaps, all_overlaps, all_redundant = [], [], [], []

for file in uploaded_files:
    supplier = os.path.splitext(file.name)[0]
    try:
        rows = load_supplier_rows(file, supplier)
        if not rows:
            st.warning(f"No valid rows found in {file.name}")
            continue
        best_map = build_best_rate_map(rows, min_amt, max_amt)
        best_segments = compress_best_map(best_map)
        supplier_maps[supplier] = best_map
        all_best_segments.extend(best_segments)
        all_gaps.extend(find_coverage_gaps(rows, min_amt, max_amt))
        all_overlaps.extend(find_overlaps(rows, min_amt, max_amt))
        all_redundant.extend(find_redundant_rules(rows, best_map))
        summary_rows.append({
            "supplier": supplier,
            "operators_found": ", ".join(sorted(set(r["operator"] for r in rows))),
            "rows_loaded": len(rows),
            "best_slabs_created": len(best_segments),
        })
    except Exception as e:
        st.error(f"Error processing {file.name}: {e}")

summary_df = pd.DataFrame(summary_rows)
best_df = pd.DataFrame(all_best_segments)
gaps_df = pd.DataFrame(all_gaps)
overlaps_df = pd.DataFrame(all_overlaps)
redundant_df = pd.DataFrame(all_redundant)
suppliers = list(supplier_maps.keys())

all_opportunities = []
if len(suppliers) >= 2:
    for a, b in combinations(suppliers, 2):
        opp = compress_opportunities(compare_suppliers(supplier_maps[a], supplier_maps[b], a, b))
        for row in opp:
            row["supplier_pair"] = f"{a} vs {b}"
        all_opportunities.extend(opp)
opportunities_df = pd.DataFrame(all_opportunities)

# =========================================================
# SUMMARY CARDS
# =========================================================
st.markdown('<div class="section-header">Overview</div>', unsafe_allow_html=True)

c1, c2, c3, c4, c5, c6 = st.columns(6)
max_profit = round(opportunities_df["profit_percent"].max(), 4) if not opportunities_df.empty else 0
_preview_club = club_deals(opportunities_df) if not opportunities_df.empty else pd.DataFrame()

cards = [
    (c1, "Suppliers",      len(suppliers),                                              "files loaded",     False),
    (c2, "Operators",      best_df["operator"].nunique() if not best_df.empty else 0,   "unique",           False),
    (c3, "Best Slabs",     len(all_best_segments),                                      "rate segments",    False),
    (c4, "Opportunities",  len(opportunities_df),                                       "raw segments",     True),
    (c5, "Clubbed Deals",  len(_preview_club),                                          "portal entries",   True),
    (c6, "Max Profit",     f"{max_profit}%",                                            "best opportunity", True),
]
for col, label, val, sub, green in cards:
    cls = "metric-card metric-green" if green else "metric-card"
    with col:
        st.markdown(f"""
        <div class="{cls}">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{val}</div>
            <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

# =========================================================
# OPPORTUNITIES
# =========================================================
st.markdown('<div class="section-header">🟢 Buy / Sell Opportunities</div>', unsafe_allow_html=True)

if not opportunities_df.empty:
    top = opportunities_df.nlargest(min(3, len(opportunities_df)), "profit_percent")
    cols = st.columns(len(top))
    for i, (_, row) in enumerate(top.iterrows()):
        with cols[i]:
            st.markdown(f"""
            <div class="opp-card">
                <div class="opp-rank">Top Opportunity #{i+1}</div>
                <div class="opp-profit">+{row['profit_percent']}%</div>
                <div class="opp-detail">
                    <b style="color:#0f172a">{row['operator']}</b><br>
                    ₹{row['amount_from']} – ₹{row['amount_to']}<br>
                    <span style="color:#16a34a">Buy from {row['buy_from']}</span><br>
                    <span style="color:#dc2626">Sell to {row['sell_to']}</span><br>
                    <span style="color:#94a3b8">Buy {row['buy_rate_percent']}% → Sell {row['sell_rate_percent']}%</span>
                </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    fig_opp = chart_profit_bar(opportunities_df)
    if fig_opp:
        st.plotly_chart(fig_opp, use_container_width=True)

    def color_profit(val):
        if val > 1:     return "background-color:#dcfce7;color:#15803d;font-weight:bold"
        elif val > 0.5: return "background-color:#f0fdf4;color:#16a34a"
        elif val > 0:   return "background-color:#f7fef9;color:#4ade80"
        return ""

    st.dataframe(
        opportunities_df.style.applymap(color_profit, subset=["profit_percent"]),
        use_container_width=True, hide_index=True
    )
else:
    st.info("No buy/sell opportunities found between uploaded suppliers.")

# =========================================================
# CLUBBED DEALS  — portal-ready, sorted ascending GP
# =========================================================
st.markdown('<div class="section-header">🔗 Clubbed Deals — Portal Ready</div>', unsafe_allow_html=True)
st.markdown(
    "<small style='color:#64748b;font-family:JetBrains Mono,monospace'>"
    "Segments that share the same <b>operator · buy_from · sell_to · buy rate · sell rate</b> are merged "
    "into one deal with comma-separated slabs. Sorted lowest GP → highest."
    "</small>",
    unsafe_allow_html=True,
)

clubbed_df = club_deals(opportunities_df) if not opportunities_df.empty else pd.DataFrame()

if not clubbed_df.empty:
    saving = len(opportunities_df) - len(clubbed_df)
    st.markdown(
        f"<div style='font-family:JetBrains Mono,monospace;font-size:0.82rem;"
        f"color:#16a34a;margin:0.4rem 0 0.8rem'>"
        f"✅ <b>{len(opportunities_df)}</b> raw segments → "
        f"<b>{len(clubbed_df)}</b> clubbed deals "
        f"<span style='color:#94a3b8'>(saves you {saving} manual portal entries)</span>"
        f"</div>",
        unsafe_allow_html=True,
    )

    def color_avg(val):
        if val > 1:     return "background-color:#dcfce7;color:#15803d;font-weight:bold"
        elif val > 0.5: return "background-color:#f0fdf4;color:#16a34a"
        elif val > 0:   return "background-color:#f7fef9;color:#4ade80"
        return ""

    st.dataframe(
        clubbed_df.style.applymap(color_avg, subset=["profit_%"]),
        use_container_width=True,
        hide_index=True,
    )
else:
    st.info("Upload 2+ supplier files to generate clubbed deals.")

# =========================================================
# CHARTS
# =========================================================
st.markdown('<div class="section-header">📊 Rate Charts</div>', unsafe_allow_html=True)
ch1, ch2 = st.columns(2)
with ch1:
    fig_bar = chart_rate_comparison(best_df, suppliers)
    if fig_bar:
        st.plotly_chart(fig_bar, use_container_width=True)
with ch2:
    fig_heat = chart_heatmap(best_df)
    if fig_heat:
        st.plotly_chart(fig_heat, use_container_width=True)
    else:
        st.info("Upload 2+ suppliers to see heatmap.")

# =========================================================
# BEST RATES
# =========================================================
st.markdown('<div class="section-header">📋 Best Rates</div>', unsafe_allow_html=True)
if not best_df.empty:
    color_map = {s: SUPPLIER_COLORS[i % len(SUPPLIER_COLORS)] for i, s in enumerate(suppliers)}

    def hl_supplier(val):
        c = color_map.get(val, "#2563eb")
        r, g, b = int(c[1:3], 16), int(c[3:5], 16), int(c[5:7], 16)
        return f"background-color:rgba({r},{g},{b},0.1);color:{c};font-weight:600"

    def hl_rate(val):
        if val >= 2:     return "color:#15803d;font-weight:bold"
        elif val >= 1:   return "color:#2563eb"
        elif val >= 0.5: return "color:#475569"
        return "color:#94a3b8"

    st.dataframe(
        best_df.style.applymap(hl_supplier, subset=["supplier"]).applymap(hl_rate, subset=["rate_percent"]),
        use_container_width=True, hide_index=True
    )

# =========================================================
# FILE SUMMARY
# =========================================================
st.markdown('<div class="section-header">📁 File Summary</div>', unsafe_allow_html=True)
if not summary_df.empty:
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

# =========================================================
# DIAGNOSTICS
# =========================================================
st.markdown('<div class="section-header">🔍 Rule Diagnostics</div>', unsafe_allow_html=True)
tab1, tab2, tab3 = st.tabs(["⚠️  Coverage Gaps", "🔁  Overlaps", "🗑️  Redundant Rules"])

with tab1:
    if not gaps_df.empty:
        st.dataframe(gaps_df, use_container_width=True, hide_index=True)
    else:
        st.success("No coverage gaps found.")
with tab2:
    if not overlaps_df.empty:
        st.dataframe(overlaps_df, use_container_width=True, hide_index=True)
    else:
        st.success("No overlapping rules found.")
with tab3:
    if not redundant_df.empty:
        st.dataframe(redundant_df, use_container_width=True, hide_index=True)
    else:
        st.success("No redundant rules found.")

# =========================================================
# EXPORTS
# =========================================================
st.markdown('<div class="section-header">⬇️ Export</div>', unsafe_allow_html=True)
dl1, dl2, dl3, dl4 = st.columns(4)
excel_bytes = to_excel_bytes(summary_df, best_df, opportunities_df, gaps_df, overlaps_df, redundant_df, clubbed_df if not clubbed_df.empty else None)

with dl1:
    st.download_button("📥 Full Analysis (Excel)", data=excel_bytes,
                       file_name="recharge_rate_analysis.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)
with dl2:
    if not best_df.empty:
        st.download_button("📥 Best Rates (CSV)", data=best_df.to_csv(index=False).encode(),
                           file_name="best_rates.csv", mime="text/csv", use_container_width=True)
with dl3:
    if not opportunities_df.empty:
        st.download_button("📥 Opportunities (CSV)", data=opportunities_df.to_csv(index=False).encode(),
                           file_name="buy_sell_opportunities.csv", mime="text/csv", use_container_width=True)
with dl4:
    if not clubbed_df.empty:
        st.download_button("📥 Clubbed Deals (CSV)", data=clubbed_df.to_csv(index=False).encode(),
                           file_name="clubbed_deals.csv", mime="text/csv", use_container_width=True)
